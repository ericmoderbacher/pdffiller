import openpyxl
import fillpdf
from fillpdf import fillpdfs


def main():
    wb = openpyxl.load_workbook('book1.xlsx')
    ws = wb.active
    test = fillpdfs.get_form_fields('test.pdf')

    for i in test:
        print(i)
        data_dict_test = {}
        data_dict_test[i] = "Yes"
        fillpdfs.write_fillable_pdf("test.pdf", "testfields/" + i + ".pdf", data_dict_test)

    fieldColDict = {}
    for col in range(ws.max_row):
        if(str(ws.cell(2, col+1).value) != 'None'):
            fieldColDict[ws.cell(2, col+1).value] = col+1
            print("field: " + str(ws.cell(2, col+1).value))
            print("col: " + str(col))
            if str(ws.cell(2, col+1).value) not in test:
                print(str(ws.cell(2, col+1).value) + "not in test!")

    for row in range(3, ws.max_row):
        print(row)
        if str(ws.cell(row, 1).value)  != 'None':
            data_dict = {}
            for field in fieldColDict:
                if (str(ws.cell(row,fieldColDict[field]).value) != 'None'):
                    data_dict[field] = ws.cell(row,fieldColDict[field]).value
            try:
                name = str("output/OutputFilename" + ws.cell(row, 1).value + ".pdf")
                print(name)
                fillpdfs.write_fillable_pdf("test.pdf", name, data_dict)
            except:
                print("Mess up")


if __name__ == '__main__':
    main()
